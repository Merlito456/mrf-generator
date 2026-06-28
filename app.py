import streamlit as st
import pandas as pd
import io
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from datetime import date
import os

st.set_page_config(page_title="MRF Generator", layout="wide")
st.title("🏗️ Automated MRF Generator with Parts Selection")

# --- DATA LOADING FUNCTIONS ---
@st.cache_data
def load_material_data():
    """Load material database from CSV"""
    try:
        mat_df = pd.read_csv("eul_material_list.csv", header=1, encoding='latin1', engine='python', on_bad_lines='skip', index_col=0)
        master_df = pd.read_csv("GLOBE SITE MASTERLIST.csv", encoding='latin1', engine='python', on_bad_lines='skip')
        master_df.rename(columns={master_df.columns[0]: 'PLAID'}, inplace=True)
        master_df.set_index('PLAID', inplace=True)
        return mat_df.join(master_df[['SITE', 'SITE_ADD']], how='left')
    except Exception as e:
        st.error(f"Error loading material data: {e}")
        return pd.DataFrame()

@st.cache_data
def load_parts_database():
    """Load parts database from Excel file with PART NUMBER, DESCRIPTION, and CATEGORY columns"""
    try:
        # Check if file exists
        if os.path.exists("PartsDatabase.xlsx"):
            parts_df = pd.read_excel("PartsDatabase.xlsx", sheet_name=0)
            
            # Map columns to expected format
            column_mapping = {}
            
            # Look for PART NUMBER column
            part_cols = [col for col in parts_df.columns if 'PART' in col.upper() and 'NUMBER' in col.upper()]
            if part_cols:
                column_mapping[part_cols[0]] = 'Part_Number'
            elif 'PART NUMBER' in parts_df.columns:
                column_mapping['PART NUMBER'] = 'Part_Number'
            elif 'PART_NO' in parts_df.columns:
                column_mapping['PART_NO'] = 'Part_Number'
            elif 'PART_NUMBER' in parts_df.columns:
                column_mapping['PART_NUMBER'] = 'Part_Number'
            elif 'PART' in parts_df.columns:
                column_mapping['PART'] = 'Part_Number'
            
            # Look for DESCRIPTION column
            desc_cols = [col for col in parts_df.columns if 'DESCRIPTION' in col.upper() or 'DESC' in col.upper()]
            if desc_cols:
                column_mapping[desc_cols[0]] = 'Part_Name'
            elif 'DESCRIPTION' in parts_df.columns:
                column_mapping['DESCRIPTION'] = 'Part_Name'
            elif 'DESC' in parts_df.columns:
                column_mapping['DESC'] = 'Part_Name'
            elif 'PART_NAME' in parts_df.columns:
                column_mapping['PART_NAME'] = 'Part_Name'
            
            # Look for CATEGORY column
            cat_cols = [col for col in parts_df.columns if 'CATEGORY' in col.upper() or 'CAT' in col.upper()]
            if cat_cols:
                column_mapping[cat_cols[0]] = 'Category'
            elif 'CATEGORY' in parts_df.columns:
                column_mapping['CATEGORY'] = 'Category'
            elif 'CAT' in parts_df.columns:
                column_mapping['CAT'] = 'Category'
            
            # Rename columns if mapping found
            if column_mapping:
                parts_df.rename(columns=column_mapping, inplace=True)
            
            # Add default columns if missing
            if 'Category' not in parts_df.columns:
                parts_df['Category'] = 'General'
            if 'Unit' not in parts_df.columns:
                parts_df['Unit'] = 'pcs'
            if 'Standard_Qty' not in parts_df.columns:
                parts_df['Standard_Qty'] = 0
            
            # Clean up - remove rows with empty part numbers
            parts_df = parts_df[parts_df['Part_Number'].notna()]
            parts_df['Part_Number'] = parts_df['Part_Number'].astype(str).str.strip()
            parts_df = parts_df[parts_df['Part_Number'] != '']
            
            # Fill any missing categories with 'General'
            parts_df['Category'] = parts_df['Category'].fillna('General')
            
            return parts_df
        else:
            st.warning("PartsDatabase.xlsx not found. Please ensure it exists in the current directory.")
            return pd.DataFrame()
    except Exception as e:
        st.error(f"Error loading PartsDatabase.xlsx: {e}")
        return pd.DataFrame()

@st.cache_data
def load_site_masterlist():
    """Load site masterlist with ability to add new sites"""
    try:
        if os.path.exists("GLOBE SITE MASTERLIST.csv"):
            df = pd.read_csv("GLOBE SITE MASTERLIST.csv", encoding='latin1', engine='python', on_bad_lines='skip')
            df.rename(columns={df.columns[0]: 'PLAID'}, inplace=True)
            return df
        else:
            # Create empty dataframe with required columns
            return pd.DataFrame(columns=['PLAID', 'SITE', 'SITE_ADD'])
    except Exception as e:
        st.error(f"Error loading site masterlist: {e}")
        return pd.DataFrame()

# Helper function to safely convert to int
def safe_int_convert(value):
    """Safely convert value to int, return 0 if conversion fails"""
    try:
        if pd.isna(value):
            return 0
        if isinstance(value, (int, float)):
            return int(value)
        if isinstance(value, str):
            # Remove any non-numeric characters except decimal point
            cleaned = ''.join(c for c in value if c.isdigit() or c == '.')
            if cleaned:
                return int(float(cleaned))
        return 0
    except:
        return 0

# --- MAIN APPLICATION ---
# Load all data
material_db = load_material_data()
parts_db = load_parts_database()
site_masterlist = load_site_masterlist()

# --- SIDEBAR - SITE SELECTION ---
st.sidebar.header("🏢 Site Selection")

# Get unique site IDs
site_ids = material_db.index.dropna().unique().tolist() if not material_db.empty else []

# Add option to add new site
site_options = site_ids + ["➕ Add New Site"] if site_ids else ["➕ Add New Site"]

selected_plaid = st.sidebar.selectbox("Select Site ID", site_options)

# Handle new site addition
if selected_plaid == "➕ Add New Site":
    st.sidebar.subheader("Add New Site")
    new_plaid = st.sidebar.text_input("New Site ID (PLAID)")
    new_site_name = st.sidebar.text_input("Site Name")
    new_site_address = st.sidebar.text_area("Site Address")
    
    if st.sidebar.button("💾 Save New Site"):
        if new_plaid and new_site_name:
            # Add to material database
            new_row = pd.DataFrame({
                'SITE': [new_site_name],
                'SITE_ADD': [new_site_address]
            }, index=[new_plaid])
            
            # Add columns for all parts with default 0
            for col in material_db.columns:
                if col not in ['SITE', 'SITE_ADD']:
                    new_row[col] = 0
            
            # Add to material_db
            material_db = pd.concat([material_db, new_row])
            st.sidebar.success(f"✅ Site {new_plaid} added successfully!")
            st.rerun()
        else:
            st.sidebar.error("Please enter Site ID and Site Name")

# Get site info if valid selection
if selected_plaid in material_db.index:
    site_info = material_db.loc[selected_plaid]
else:
    site_info = pd.Series({'SITE': '', 'SITE_ADD': ''})

# --- PARTS SELECTION SECTION ---
st.header("🔧 Parts Selection and Quantities")

# Display parts database info
if not parts_db.empty:
    # Show category statistics
    if 'Category' in parts_db.columns:
        category_counts = parts_db['Category'].value_counts()
        st.info(f"📊 Loaded {len(parts_db)} parts from PartsDatabase.xlsx across {len(category_counts)} categories")
        
        # Show category distribution in expander
        with st.expander("📊 Category Distribution"):
            col_cat1, col_cat2 = st.columns(2)
            with col_cat1:
                st.dataframe(category_counts.reset_index().rename(columns={'index': 'Category', 'Category': 'Count'}), 
                           use_container_width=True, hide_index=True)
            with col_cat2:
                st.bar_chart(category_counts)
    else:
        st.info(f"📊 Loaded {len(parts_db)} parts from PartsDatabase.xlsx")
    
    # Preview first few rows
    with st.expander("📋 Preview Parts Database"):
        preview_cols = ['Part_Number', 'Part_Name']
        if 'Category' in parts_db.columns:
            preview_cols.append('Category')
        st.dataframe(parts_db[preview_cols].head(10), use_container_width=True)
        if len(parts_db) > 10:
            st.caption(f"... and {len(parts_db) - 10} more parts")
else:
    st.warning("⚠️ PartsDatabase.xlsx not loaded. You can still add custom materials manually.")

# Tabbed interface for parts selection
tab1, tab2 = st.tabs(["📦 Select from Database", "✏️ Add Custom Materials"])

# Initialize selected parts dictionary
if 'selected_parts' not in st.session_state:
    st.session_state.selected_parts = {}

# --- TAB 1: Select from Database ---
with tab1:
    if not parts_db.empty:
        # Two columns for parts selection
        col_parts1, col_parts2 = st.columns(2)
        
        with col_parts1:
            # Category filter
            if 'Category' in parts_db.columns:
                categories = ["All"] + sorted(parts_db['Category'].dropna().unique().tolist())
                selected_category = st.selectbox("🔍 Filter by Category", categories, key="category_filter")
                
                if selected_category != "All":
                    filtered_parts = parts_db[parts_db['Category'] == selected_category]
                else:
                    filtered_parts = parts_db
            else:
                filtered_parts = parts_db
                st.info("No category column found. Showing all parts.")

        with col_parts2:
            # Search functionality
            search_term = st.text_input("🔎 Search Parts", placeholder="Enter part number or description...", key="search_parts")
            if search_term and not filtered_parts.empty:
                filtered_parts = filtered_parts[
                    filtered_parts['Part_Number'].astype(str).str.contains(search_term, case=False, na=False) |
                    filtered_parts['Part_Name'].astype(str).str.contains(search_term, case=False, na=False)
                ]
        
        # Display parts with quantity selection
        if not filtered_parts.empty:
            st.subheader("📦 Select Parts and Specify Quantities")
            st.caption(f"Showing {len(filtered_parts)} parts")
            
            # Check if we have existing quantities from material_db
            existing_parts = {}
            if selected_plaid in material_db.index:
                for col in material_db.columns:
                    if col not in ['SITE', 'SITE_ADD']:
                        val = material_db.loc[selected_plaid, col]
                        if pd.notna(val) and str(val).strip() != '' and str(val) != '0':
                            existing_parts[col] = safe_int_convert(val)
            
            # Create selection table with better layout
            st.markdown("### Select parts and enter quantities")
            
            # Use columns for better layout
            col_labels = st.columns([2, 3, 1, 1, 1.5])
            with col_labels[0]:
                st.write("**Part Number**")
            with col_labels[1]:
                st.write("**Description**")
            with col_labels[2]:
                st.write("**Category**")
            with col_labels[3]:
                st.write("**Unit**")
            with col_labels[4]:
                st.write("**Quantity**")
            
            st.divider()
            
            for idx, row in filtered_parts.iterrows():
                part_num = str(row['Part_Number'])
                if pd.isna(part_num) or part_num == 'nan' or part_num == '':
                    continue
                    
                cols = st.columns([2, 3, 1, 1, 1.5])
                
                with cols[0]:
                    st.code(part_num, language=None)
                
                with cols[1]:
                    part_name = str(row.get('Part_Name', ''))
                    st.write(part_name)
                
                with cols[2]:
                    category = str(row.get('Category', 'General'))
                    st.write(category)
                
                with cols[3]:
                    unit = str(row.get('Unit', 'pcs'))
                    st.write(unit)
                
                with cols[4]:
                    # Safely get default quantity
                    default_qty = 0
                    
                    # Check existing parts from material_db
                    if part_num in existing_parts:
                        default_qty = existing_parts[part_num]
                    
                    # Check Standard_Qty if no existing quantity
                    if default_qty == 0 and 'Standard_Qty' in row:
                        default_qty = safe_int_convert(row['Standard_Qty'])
                    
                    # Check if already in session state (user's selection)
                    if part_num in st.session_state.selected_parts:
                        default_qty = st.session_state.selected_parts[part_num]
                    
                    # Ensure default_qty is an integer
                    default_qty = safe_int_convert(default_qty)
                    
                    qty = st.number_input(
                        f"qty_{part_num}",
                        min_value=0,
                        max_value=9999,
                        value=default_qty,
                        step=1,
                        key=f"db_qty_{part_num}",
                        label_visibility="collapsed"
                    )
                    if qty > 0:
                        st.session_state.selected_parts[part_num] = qty
                    elif part_num in st.session_state.selected_parts:
                        del st.session_state.selected_parts[part_num]
                
                st.divider()
        else:
            st.warning("No parts match your filters. Try adjusting your search or category filter.")
    else:
        st.info("📝 No parts database loaded. Use the 'Add Custom Materials' tab to add parts manually.")

# --- TAB 2: Add Custom Materials ---
with tab2:
    st.subheader("✏️ Add Custom Materials Not in Database")
    st.caption("Add any additional materials that are not in the PartsDatabase.xlsx")
    
    # Input for custom material
    col_custom1, col_custom2 = st.columns([1, 1])
    
    with col_custom1:
        custom_part_num = st.text_input("Part Number", placeholder="e.g., CUSTOM-001", key="custom_part_num")
        custom_description = st.text_input("Description", placeholder="e.g., Special Cable Assembly", key="custom_desc")
        custom_category = st.text_input("Category (optional)", placeholder="e.g., Accessories", key="custom_category")
    
    with col_custom2:
        custom_unit = st.selectbox("Unit", ["pcs", "m", "kg", "L", "box", "set", "roll", "pair"], key="custom_unit")
        custom_qty = st.number_input("Quantity", min_value=1, max_value=9999, value=1, step=1, key="custom_qty")
    
    # Add custom material button
    col_btn1, col_btn2 = st.columns([1, 3])
    with col_btn1:
        if st.button("➕ Add Custom Material", use_container_width=True):
            if custom_part_num and custom_description and custom_qty > 0:
                # Add to selected parts
                st.session_state.selected_parts[custom_part_num] = custom_qty
                
                # Also store description and category for display
                if 'custom_descriptions' not in st.session_state:
                    st.session_state.custom_descriptions = {}
                st.session_state.custom_descriptions[custom_part_num] = {
                    'description': custom_description,
                    'unit': custom_unit,
                    'category': custom_category if custom_category else 'Custom'
                }
                
                st.success(f"✅ Added {custom_part_num} - {custom_description} (Qty: {custom_qty})")
                
                # Clear inputs
                st.session_state.custom_part_num = ""
                st.session_state.custom_desc = ""
                st.session_state.custom_category = ""
                st.session_state.custom_qty = 1
                
                st.rerun()
            else:
                st.error("❌ Please fill in all required fields (Part Number, Description, and Quantity)")
    
    # Quick add multiple materials
    with st.expander("📋 Add Multiple Materials at Once"):
        st.caption("Enter multiple materials in CSV format: PartNumber,Description,Category,Quantity,Unit")
        st.caption("Example: CBL-001,Fiber Cable,Cables,5,m")
        st.caption("Note: Category is optional")
        
        multi_input = st.text_area(
            "Enter materials (one per line)",
            placeholder="CBL-001,Fiber Optic Cable,Cables,10,m\nBRK-002,Mounting Bracket,Hardware,5,pcs\nPSU-003,Power Supply,Electronics,2,pcs",
            height=150
        )
        
        if st.button("📥 Import Multiple Materials"):
            if multi_input:
                lines = multi_input.strip().split('\n')
                added_count = 0
                for line in lines:
                    parts = [p.strip() for p in line.split(',')]
                    if len(parts) >= 3:
                        part_num = parts[0]
                        desc = parts[1]
                        try:
                            # Check if category is provided (4th field) or quantity is 3rd
                            if len(parts) >= 4:
                                # Try to parse as PartNum,Desc,Category,Qty,Unit
                                try:
                                    qty = int(parts[3])
                                    category = parts[2]
                                    unit = parts[4] if len(parts) > 4 else 'pcs'
                                except ValueError:
                                    # If 3rd field is not a number, treat as category
                                    category = parts[2]
                                    qty = int(parts[3])
                                    unit = parts[4] if len(parts) > 4 else 'pcs'
                            else:
                                # PartNum,Desc,Qty,Unit
                                qty = int(parts[2])
                                category = 'Custom'
                                unit = parts[3] if len(parts) > 3 else 'pcs'
                            
                            if part_num and desc and qty > 0:
                                st.session_state.selected_parts[part_num] = qty
                                if 'custom_descriptions' not in st.session_state:
                                    st.session_state.custom_descriptions = {}
                                st.session_state.custom_descriptions[part_num] = {
                                    'description': desc,
                                    'unit': unit,
                                    'category': category
                                }
                                added_count += 1
                        except (ValueError, IndexError) as e:
                            st.warning(f"Skipping invalid line: {line} - {str(e)}")
                    else:
                        st.warning(f"Skipping invalid format: {line} - needs at least PartNumber, Description, Quantity")
                
                if added_count > 0:
                    st.success(f"✅ Added {added_count} custom materials successfully!")
                    st.rerun()

# --- DISPLAY SELECTED PARTS SUMMARY ---
st.header("📋 Current Selection Summary")

# Show all selected parts including custom ones
if st.session_state.selected_parts:
    summary_data = []
    
    for part_num, qty in st.session_state.selected_parts.items():
        # Check if it's from database
        if not parts_db.empty:
            part_row = parts_db[parts_db['Part_Number'] == part_num]
            if not part_row.empty:
                part_row = part_row.iloc[0]
                summary_data.append({
                    'Part Number': part_num,
                    'Description': part_row.get('Part_Name', ''),
                    'Category': part_row.get('Category', 'General'),
                    'Quantity': qty,
                    'Unit': part_row.get('Unit', 'pcs'),
                    'Source': 'Database'
                })
                continue
        
        # Check if it's custom
        if 'custom_descriptions' in st.session_state and part_num in st.session_state.custom_descriptions:
            custom_info = st.session_state.custom_descriptions[part_num]
            summary_data.append({
                'Part Number': part_num,
                'Description': custom_info.get('description', ''),
                'Category': custom_info.get('category', 'Custom'),
                'Quantity': qty,
                'Unit': custom_info.get('unit', 'pcs'),
                'Source': 'Custom'
            })
        else:
            # Fallback for custom without description
            summary_data.append({
                'Part Number': part_num,
                'Description': 'Custom Material',
                'Category': 'Custom',
                'Quantity': qty,
                'Unit': 'pcs',
                'Source': 'Custom'
            })
    
    if summary_data:
        summary_df = pd.DataFrame(summary_data)
        st.dataframe(summary_df, use_container_width=True, hide_index=True)
        
        # Statistics and actions
        col_stat1, col_stat2, col_stat3, col_stat4 = st.columns([1, 1, 1, 1])
        with col_stat1:
            total_parts = len(st.session_state.selected_parts)
            total_qty = sum(st.session_state.selected_parts.values())
            st.info(f"📦 {total_parts} part types, {total_qty} total quantity")
        
        with col_stat2:
            db_count = len([p for p in summary_data if p['Source'] == 'Database'])
            custom_count = len([p for p in summary_data if p['Source'] == 'Custom'])
            st.info(f"📊 {db_count} from database, {custom_count} custom")
        
        with col_stat3:
            # Show category breakdown
            if 'Category' in summary_df.columns:
                categories = summary_df['Category'].value_counts()
                st.info(f"📂 {len(categories)} categories")
        
        with col_stat4:
            if st.button("🗑️ Clear All Selections", use_container_width=True):
                st.session_state.selected_parts = {}
                if 'custom_descriptions' in st.session_state:
                    st.session_state.custom_descriptions = {}
                st.rerun()
else:
    st.info("ℹ️ No parts selected yet. Please select parts from the database or add custom materials.")

# --- USER INPUTS ---
st.header("📝 MRF Details")
col1, col2 = st.columns(2)

with col1:
    destination_city = st.text_input("📍 Enter Destination City")
    received_by = st.text_input("👤 Received By")
    mod_site_add = st.text_area(
        "🏠 Site Address", 
        value=str(site_info.get('SITE_ADD', '')) if isinstance(site_info, pd.Series) else ''
    )

with col2:
    uploaded_file = st.file_uploader("✍️ Upload E-Signature", type=['png', 'jpg', 'jpeg'])
    if uploaded_file:
        st.image(uploaded_file, caption="Signature Preview", width=200)

# --- GENERATE MRF BUTTON ---
st.divider()
col_gen1, col_gen2, col_gen3 = st.columns([1, 2, 1])
with col_gen2:
    generate_button = st.button("🚀 Generate MRF", type="primary", use_container_width=True)

if generate_button:
    if not selected_plaid or selected_plaid == "➕ Add New Site":
        st.error("❌ Please select a valid site.")
    elif not destination_city:
        st.error("❌ Please enter Destination City.")
    elif not received_by:
        st.error("❌ Please enter Received By.")
    elif not st.session_state.selected_parts:
        st.warning("⚠️ No parts selected. Please select at least one part with quantity > 0.")
    else:
        try:
            with st.spinner("Generating MRF..."):
                # Load template
                wb = load_workbook("template_mrf.xlsx")
                ws = wb.active
                
                # 1. Replace Text Placeholders
                replacements = {
                    "[CITY]": destination_city,
                    "[PLAID  - SITE NAME]": f"{selected_plaid} - {site_info.get('SITE', 'Unknown Site')}",
                    "[SITE_ADD]": mod_site_add,
                    "[RECEIVED BY]": received_by,
                    "[DATE_GENERATED]": date.today().strftime("%Y-%m-%d"),
                    "[ESIG]": ""
                }
                
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str):
                            for placeholder, value in replacements.items():
                                if placeholder in cell.value:
                                    cell.value = cell.value.replace(placeholder, value)
                
                # 2. Update Parts Quantities
                current_row = 16
                
                # Clear existing quantities in rows 16-60
                for row in range(16, 61):
                    ws[f'D{row}'] = ""
                
                # Add selected parts (both from database and custom)
                for part_num, qty in st.session_state.selected_parts.items():
                    if qty > 0:
                        # Ensure qty is integer
                        qty = safe_int_convert(qty)
                        
                        # Check if part exists in template (rows 16-60)
                        found = False
                        for row in range(16, 61):
                            part_cell = ws[f'A{row}']
                            if part_cell.value and str(part_cell.value).strip() == part_num:
                                ws[f'D{row}'] = qty
                                found = True
                                break
                        
                        # If not found, add to new row
                        if not found:
                            current_row += 1
                            ws[f'A{current_row}'] = part_num
                            ws[f'D{current_row}'] = qty
                            
                            # Add description if available (custom)
                            if 'custom_descriptions' in st.session_state and part_num in st.session_state.custom_descriptions:
                                # If there's a column for description in template
                                desc_cell = ws[f'B{current_row}']
                                if desc_cell and isinstance(desc_cell.value, str) or desc_cell.value is None:
                                    ws[f'B{current_row}'] = st.session_state.custom_descriptions[part_num].get('description', '')
                
                # 3. Add Additional Parts from Material DB (not selected manually)
                if selected_plaid in material_db.index:
                    for col in material_db.columns:
                        if col not in ['SITE', 'SITE_ADD'] and col not in st.session_state.selected_parts:
                            val = material_db.loc[selected_plaid, col]
                            if pd.notna(val) and str(val).strip() != '' and str(val) != '0':
                                current_row += 1
                                ws[f'A{current_row}'] = col
                                ws[f'D{current_row}'] = safe_int_convert(val)
                
                # 4. Inject Signature Image
                if uploaded_file:
                    try:
                        target_cell = 'D47'
                        
                        # Check if D47 is part of a merge
                        for merged_range in ws.merged_cells.ranges:
                            if 'D47' in merged_range:
                                target_cell = merged_range.start_cell.coordinate
                                break
                        
                        ws[target_cell] = ""
                        
                        img = ExcelImage(uploaded_file)
                        img.width = 120
                        img.height = 60
                        ws.add_image(img, target_cell)
                        
                    except Exception as e:
                        st.warning(f"⚠️ Could not place signature: {e}")
                
                # Save and Trigger Download
                output = io.BytesIO()
                wb.save(output)
                filename = f"NOKIA-MRF_{selected_plaid}-{site_info.get('SITE', 'Unknown')}_{date.today().strftime('%Y%m%d')}.xlsx"
                
                st.success("✅ MRF Generated Successfully!")
                st.balloons()
                st.download_button(
                    "📥 Download Final MRF", 
                    data=output.getvalue(), 
                    file_name=filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
                
        except Exception as e:
            st.error(f"❌ Error generating MRF: {e}")
            st.exception(e)

# --- FOOTER ---
st.divider()
st.caption("ℹ️ **PartsDatabase.xlsx** should have columns: **PART NUMBER**, **DESCRIPTION**, and **CATEGORY**")
st.caption("📌 Category column helps organize parts. You can also add custom materials not in the database.")